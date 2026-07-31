"""Esportazione del calendario in formato Excel (.xlsx), in aggiunta alla
stampa/PDF della Fase 6: stessa logica di lettura dei dati (_dati_calendario_sede),
un foglio per sede."""

import re
from calendar import monthrange
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.auth import RUOLI_LETTURA, richiedi_ruolo
from app.database import get_db
from app.models import Sede, Utente
from app.routers.calendario import (
    _anno_mese_validi_o_oggi,
    _dati_calendario_sede,
    _giorni_del_mese,
)

router = APIRouter()

COLORE_INTESTAZIONE = "1F2430"
COLORE_WEEKEND = "FDF1E0"
COLORE_ASSENZA = "D64545"

_CARATTERI_NON_AMMESSI_FOGLIO = re.compile(r'[:\\/?*\[\]]')


def _titolo_foglio(nome: str, gia_usati: set[str]) -> str:
    """openpyxl rifiuta alcuni caratteri nel titolo di un foglio Excel
    (: \\ / ? * [ ]) sollevando un ValueError non gestito: Sede.nome è
    testo libero senza questa restrizione, e un nome come "Via Roma 5/A"
    (tutt'altro che raro per un indirizzo) fa fallire con un 500 l'intera
    esportazione invece di produrre il file. Deduplica anche il caso in
    cui due sedi, una volta sanificate e troncate ai 31 caratteri massimi
    di Excel, finiscano con lo stesso titolo: create_sheet fallirebbe
    comunque su un duplicato."""
    pulito = _CARATTERI_NON_AMMESSI_FOGLIO.sub(" ", nome).strip() or "Sede"
    titolo = pulito[:31]
    contatore = 2
    while titolo in gia_usati:
        suffisso = f" ({contatore})"
        titolo = pulito[:31 - len(suffisso)] + suffisso
        contatore += 1
    gia_usati.add(titolo)
    return titolo


def _testo_cella(assegnazione, sostituzioni_giorno) -> str:
    sostituzione_intera = next((s for s in sostituzioni_giorno if s.ora_inizio is None), None)
    if sostituzione_intera:
        sostituto = sostituzione_intera.dipendente_sostituto
        return f"SOST: {sostituto.cognome} {sostituto.nome}"

    if assegnazione and assegnazione.origine == "assenza":
        base = "ASSENTE"
    elif assegnazione and assegnazione.tipo_turno:
        base = assegnazione.tipo_turno.etichetta
    else:
        base = ""

    # Sostituzioni orarie (solo una fascia della giornata, vedi
    # Sostituzione.ora_inizio/ora_fine): a schermo restano visibili come
    # badge accanto al turno del titolare (_cella_calendario.html), quindi
    # anche qui vanno aggiunte invece di sparire silenziosamente.
    sostituzioni_orarie = [s for s in sostituzioni_giorno if s.ora_inizio is not None]
    if sostituzioni_orarie:
        dettagli = "; ".join(
            f"SOST {s.ora_inizio.strftime('%H:%M')}-{s.ora_fine.strftime('%H:%M')}: "
            f"{s.dipendente_sostituto.cognome} {s.dipendente_sostituto.nome}"
            for s in sostituzioni_orarie
        )
        return f"{base} ({dettagli})" if base else dettagli

    return base


@router.get("/calendario/excel")
def esporta_excel(
    sede_id: int | None = None,
    tutte: bool = False,
    anno: int | None = None,
    mese: int | None = None,
    db: Session = Depends(get_db),
    utente: Utente = Depends(richiedi_ruolo(*RUOLI_LETTURA)),
):
    anno, mese = _anno_mese_validi_o_oggi(anno, mese)
    numero_giorni = monthrange(anno, mese)[1]
    giorni = _giorni_del_mese(anno, mese)

    if tutte:
        sedi_da_esportare = db.query(Sede).filter(Sede.attivo == True).order_by(Sede.nome).all()  # noqa: E712
    else:
        sede = (
            db.query(Sede).filter(Sede.id == sede_id, Sede.attivo == True).first()  # noqa: E712
            if sede_id
            else db.query(Sede).filter(Sede.attivo == True).order_by(Sede.nome).first()  # noqa: E712
        )
        sedi_da_esportare = [sede] if sede else []

    cartella_lavoro = Workbook()
    cartella_lavoro.remove(cartella_lavoro.active)
    titoli_foglio_usati: set[str] = set()

    for sede in sedi_da_esportare:
        dipendenti, assegnazioni_per_dipendente, sostituzioni_per_dipendente, _, _ = _dati_calendario_sede(
            db, sede, anno, mese, numero_giorni
        )
        foglio = cartella_lavoro.create_sheet(title=_titolo_foglio(sede.nome, titoli_foglio_usati))

        intestazione = foglio.cell(row=1, column=1, value="Dipendente")
        intestazione.font = Font(bold=True, color="FFFFFF")
        intestazione.fill = PatternFill("solid", fgColor=COLORE_INTESTAZIONE)
        for indice, giorno in enumerate(giorni, start=2):
            cella = foglio.cell(row=1, column=indice, value=f"{giorno['numero']} {giorno['iniziale']}")
            cella.font = Font(bold=True, color="FFFFFF")
            cella.fill = PatternFill("solid", fgColor=COLORE_INTESTAZIONE)
            cella.alignment = Alignment(horizontal="center")

        for riga_indice, dipendente in enumerate(dipendenti, start=2):
            foglio.cell(row=riga_indice, column=1, value=f"{dipendente.cognome} {dipendente.nome}")
            for colonna_indice, giorno in enumerate(giorni, start=2):
                assegnazione = assegnazioni_per_dipendente.get(dipendente.id, {}).get(giorno["numero"])
                sostituzioni_giorno = sostituzioni_per_dipendente.get(dipendente.id, {}).get(giorno["numero"], [])
                testo = _testo_cella(assegnazione, sostituzioni_giorno)
                cella = foglio.cell(row=riga_indice, column=colonna_indice, value=testo)
                cella.alignment = Alignment(horizontal="center")
                if giorno["weekend"]:
                    cella.fill = PatternFill("solid", fgColor=COLORE_WEEKEND)
                if testo == "ASSENTE":
                    cella.font = Font(color=COLORE_ASSENZA, bold=True)

        foglio.column_dimensions["A"].width = 22
        for indice in range(2, len(giorni) + 2):
            foglio.column_dimensions[get_column_letter(indice)].width = 10
        foglio.freeze_panes = "B2"

    if not cartella_lavoro.sheetnames:
        cartella_lavoro.create_sheet(title="Nessuna sede")

    buffer = BytesIO()
    cartella_lavoro.save(buffer)
    buffer.seek(0)

    nome_file = f"calendario_{anno}-{mese:02d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_file}"'},
    )
