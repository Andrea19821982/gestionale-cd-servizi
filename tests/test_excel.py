from datetime import date, time

from io import BytesIO

from openpyxl import load_workbook

from app.models import AssegnazioneGiornaliera, Dipendente, Sede, Sostituzione, TipoTurno
from tests.conftest import login


def _crea_sede(db, nome="Sede Test", colore="#123456"):
    sede = Sede(nome=nome, colore_hex=colore, attivo=True)
    db.add(sede)
    db.commit()
    db.refresh(sede)
    return sede


def _login_admin(client, crea_utente):
    crea_utente("admin_test", "passwordsegreta", "amministratore")
    login(client, "admin_test", "passwordsegreta")


def test_excel_singola_sede(client, crea_utente, db):
    _login_admin(client, crea_utente)
    sede = _crea_sede(db, "Sede Excel Singola")
    dip = Dipendente(cognome="Excel", nome="Test", sede_riferimento_id=sede.id, attivo=True)
    db.add(dip)
    db.commit()

    r = client.get(f"/calendario/excel?sede_id={sede.id}&anno=2026&mese=8")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    cartella = load_workbook(BytesIO(r.content))
    assert cartella.sheetnames == ["Sede Excel Singola"]
    foglio = cartella["Sede Excel Singola"]
    assert foglio.cell(row=1, column=1).value == "Dipendente"
    assert foglio.cell(row=2, column=1).value == "Excel Test"


def test_excel_tutte_le_sedi(client, crea_utente, db):
    _login_admin(client, crea_utente)
    _crea_sede(db, "Excel Sede A")
    _crea_sede(db, "Excel Sede B")

    r = client.get("/calendario/excel?tutte=1&anno=2026&mese=8")
    assert r.status_code == 200
    cartella = load_workbook(BytesIO(r.content))
    assert "Excel Sede A" in cartella.sheetnames
    assert "Excel Sede B" in cartella.sheetnames


def test_excel_sede_inesistente_non_crasha(client, crea_utente):
    _login_admin(client, crea_utente)
    r = client.get("/calendario/excel?sede_id=9999&anno=2026&mese=8")
    assert r.status_code == 200
    cartella = load_workbook(BytesIO(r.content))
    assert cartella.sheetnames == ["Nessuna sede"]


def test_excel_sede_con_caratteri_non_ammessi_nel_titolo_foglio(client, crea_utente, db):
    """openpyxl rifiuta : \\ / ? * [ ] nel titolo di un foglio: un nome
    sede realistico come un indirizzo con lo slash non deve far fallire
    l'intera esportazione con un 500."""
    _login_admin(client, crea_utente)
    sede = _crea_sede(db, "Via Roma 5/A")
    dip = Dipendente(cognome="Slash", nome="Test", sede_riferimento_id=sede.id, attivo=True)
    db.add(dip)
    db.commit()

    r = client.get(f"/calendario/excel?sede_id={sede.id}&anno=2026&mese=8")
    assert r.status_code == 200
    cartella = load_workbook(BytesIO(r.content))
    assert cartella.sheetnames == ["Via Roma 5 A"]


def test_excel_due_sedi_con_lo_stesso_titolo_foglio_dopo_la_sanificazione(client, crea_utente, db):
    """Due sedi che, una volta tolti i caratteri non ammessi, finiscono
    con lo stesso titolo non devono far fallire create_sheet su un
    duplicato: la seconda va distinta con un suffisso."""
    _login_admin(client, crea_utente)
    _crea_sede(db, "Sede/Duplicata")
    _crea_sede(db, "Sede:Duplicata")

    r = client.get("/calendario/excel?tutte=1&anno=2026&mese=8")
    assert r.status_code == 200
    cartella = load_workbook(BytesIO(r.content))
    assert cartella.sheetnames == ["Sede Duplicata", "Sede Duplicata (2)"]


def test_excel_richiede_login(client):
    r = client.get("/calendario/excel", follow_redirects=False)
    assert r.status_code == 303
    assert r.headers["location"].startswith("/login")


def test_excel_accessibile_a_consultazione(client, crea_utente, db):
    crea_utente("consultazione_test", "passwordsegreta", "consultazione")
    login(client, "consultazione_test", "passwordsegreta")
    sede = _crea_sede(db, "Sede Excel Consultazione")

    r = client.get(f"/calendario/excel?sede_id={sede.id}&anno=2026&mese=8")
    assert r.status_code == 200


def test_excel_mostra_la_sostituzione_oraria_come_nel_calendario(client, crea_utente, db):
    """Nel calendario a schermo (vedi templates/_cella_calendario.html) una
    sostituzione oraria (solo una fascia della giornata, Sostituzione con
    ora_inizio/ora_fine valorizzati) resta visibile come badge accanto al
    turno del titolare: non sparisce. L'export Excel invece, prima del fix,
    ignorava del tutto le sostituzioni orarie in _testo_cella (controllava
    solo quella a giornata intera, ora_inizio is None) e mostrava la cella
    come se il titolare avesse lavorato il suo turno normale senza alcuna
    sostituzione: un disallineamento silenzioso tra schermo ed export."""
    _login_admin(client, crea_utente)
    sede = _crea_sede(db, "Sede Excel Oraria")
    tipo = TipoTurno(etichetta="Mattina Excel Oraria", ora_inizio=time(7, 0), ora_fine=time(13, 30))
    db.add(tipo)
    db.commit()
    db.refresh(tipo)

    titolare = Dipendente(cognome="Titolare", nome="Excel", sede_riferimento_id=sede.id, attivo=True)
    sostituto = Dipendente(cognome="Sostituto", nome="Excel", sede_riferimento_id=sede.id, attivo=True)
    db.add_all([titolare, sostituto])
    db.commit()
    for d in (titolare, sostituto):
        db.refresh(d)

    db.add(AssegnazioneGiornaliera(
        dipendente_id=titolare.id, data=date(2026, 8, 10), sede_effettiva_id=sede.id,
        tipo_turno_id=tipo.id, origine="manuale",
    ))
    db.add(Sostituzione(
        data=date(2026, 8, 10), dipendente_partente_id=titolare.id, sede_partenza_id=sede.id,
        dipendente_sostituto_id=sostituto.id, sede_arrivo_id=sede.id,
        ora_inizio=time(9, 0), ora_fine=time(11, 0),
    ))
    db.commit()

    r = client.get(f"/calendario/excel?sede_id={sede.id}&anno=2026&mese=8")
    assert r.status_code == 200
    cartella = load_workbook(BytesIO(r.content))
    foglio = cartella["Sede Excel Oraria"]

    riga_titolare = None
    for riga in range(2, foglio.max_row + 1):
        if foglio.cell(row=riga, column=1).value == "Titolare Excel":
            riga_titolare = riga
            break
    assert riga_titolare is not None

    testo_10_agosto = foglio.cell(row=riga_titolare, column=1 + 10).value
    assert "Sostituto" in testo_10_agosto or "09:00" in testo_10_agosto
